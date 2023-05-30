from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from random import *

# load_workbook() 함수는 파일을 불러올 수 있다.
#
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

print(ws["A2"]) # 셀에 대한 정보 출력


wb.close()