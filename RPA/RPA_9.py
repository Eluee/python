from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
import datetime

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("testsampleb.xlsx", data_only = True)
ws = wb.active

for row in ws.values:
    for cell in row:
        print(cell)

# 수식을 가져옴



wb.save("testsampleb.xlsx")
wb.close()