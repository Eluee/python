from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
import datetime

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = Workbook()
ws = wb.active

ws.merge_cells("B2:D2")
ws["B2"].value = "Merged Cell"

wb.save("testsample3.xlsx")
wb.close()