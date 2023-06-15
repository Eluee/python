from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from random import *

# load_workbook() 함수는 파일을 불러올 수 있다.
#
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

# iter_rows()함수의 매개변수 값이 min_row = 2 면 2번 행을 포함한다. 즉 min_row = 1만 있다면 결국 한줄을 포함하기 때문에 전체 행을 말한다.

for row in ws.iter_rows(min_row = 2):
    if int(row[1].value) > 50:
        print(row[0].value, "번 학생은 50점을 넘었습니다.")

wb.close()