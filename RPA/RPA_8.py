from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
import datetime

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] ="=SUM(1, 2, 3)" # 1 + 2 + 3 = 6 합계
ws["A3"] ="=AVERAGE(1, 2, 3)" #2 (평균)

ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)" #30
ws["A7"] = "=COUNT(A2:A6)"

wb.save("testsampleb.xlsx")
wb.close()