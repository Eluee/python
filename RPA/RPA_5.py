from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from random import *

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

# 해당 영역을 행을 아래로 5칸 이동 열을 왼쪽으로 1칸 이동
ws.move_range("C1:C11", rows=5, cols=-1) 

wb.save("testsampleb.xlsx")
wb.close()