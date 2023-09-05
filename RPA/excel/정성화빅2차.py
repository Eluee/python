from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("정성화빅2차.xlsx")
ws = wb.active
table = [list(row) + [0, 0] for row in ws.values]

# 2. H열에 총점 (sum 이용), I열에 성적정보 추가
table[0][len(table[0]) - 1] = "성적정보"
table[0][len(table[0]) - 2] = "총점"

grades = [(0, "D"),(70, "C"),(80, "B"),(90,"A")]
goodtable = [["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사","프로젝트", "총점","성적정보"]]

for x in range(1, len(table)):
    # 1. 퀴즈2 점수를 10으로 수정
    table[x][3] = 10
    #결석 횟수 출석점수로 수정 (결석 1회당 1점 감점)
    table[x][1] = 10 - table[x][1]
    score = sum(table[x][1:])
    table[x][len(table[0]) - 2] = "=SUM(B"+ str(x + 1) +":G" + str(x + 1) + ")"
    if table[x][1] <= 6:
        # 결석 4회 이상인 학생은 총점 상관없이 F
        table[x][len(table[0]) - 1] = "F"
    else:
        # ~ 총점 90이상 A, 80이상 B, 70 이사 C 나머지 D
        for grade in grades:
            if score >= grade[0]:
                table[x][len(table[0]) - 1] = grade[1]
        if table[x][len(table[0]) - 1] == "A":
            goodtable.append(table[x])
            goodtable[len(goodtable) - 1][len(goodtable[0]) - 2] = "=SUM(B"+ str(len(goodtable)) +":G" + str(len(goodtable)) + ")"
                 
for row in range(len(table)):
   for col in range(len(table[0])):
        ws.cell(row + 1, col + 1, table[row][col])
        
wb.create_sheet("good")
wb.save("정성화빅2차.xlsx")
ws = wb['good']
for row in goodtable:
    ws.append(row)

wb.save("정성화빅2차.xlsx")
wb.close()

