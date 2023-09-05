from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from random import *

# load_workbook() 함수는 파일을 불러올 수 있다.
#
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

ws.insert_rows(8) #8번째 줄에 비어있는 한 줄 추가
ws.insert_cols(2) #2번째 열에 빈 한 줄 추가

ws.delete_rows(8) #8번째 행을 제거
ws.delete_cols(2) #2번쨰 열을 제거

wb.save("testsampleb.xlsx")
wb.close()