from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = Workbook()
ws = wb.active
table = [
    ["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로젝트"],
    [1,3,8,5,14,26,12],
    [2,2,3,7,15,24,18],
    [3,1,5,8,8,12,4],
    [4,0,8,7,17,21,18],
    [5,4,8,7,16,25,15],
    [6,0,5,8,8,17,0],
    [7,0,9,10,16,27,18],
    [8,4,6,6,15,19,17],
    [9,0,10,9,19,30,19],
    [10,0,8,8,20,25,20]]

for row in table:
    ws.append(row)

wb.save("정성화빅2차.xlsx")
wb.close()