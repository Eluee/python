from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from openpyxl.drawing.image import Image

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = Workbook()
ws = wb.active

img = Image('img/pepe.png')
ws.add_image(img, "C3")

wb.save("testsample_image.xlsx")
wb.close()