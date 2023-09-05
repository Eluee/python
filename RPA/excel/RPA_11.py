from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("testsample3.xlsx")
ws = wb.active

ws.unmerge_cells("B2:D2")
wb.save("testsample3.xlsx")
wb.close()