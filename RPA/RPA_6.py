from openpyxl import Workbook
from openpyxl import load_workbook # 워크북 
from random import *

from openpyxl.chart import BarChart, Reference, LineChart

# load_workbook() 함수는 파일을 불러올 수 있다.
wb = load_workbook("testsampleb.xlsx")
ws = wb.active

#B2:C11 까지의 데이터를 차트로 생성
bar_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col = 3)
bar_chart = BarChart()
bar_chart.add_data(bar_value, titles_from_data= True)
ws.add_chart(bar_chart, "E15")



line_value = Reference(ws, min_row=1, max_row=11, min_col=2, max_col=3)
# Reference() 참조 함수로 해당 영역의 데이터를 참조한다.

line_chart = LineChart()
# LineChart() 함수를 통해 Line 형식의 chart의 객체를 가져옴
line_chart.add_data(line_value, titles_from_data=True)
# LineChart.add_data() 메소드로 chart에 대한 데이터와 속성을 지정 
# titles_from_data=True 데이터에 대한 타이틀을 가져와서 표시하는지에 대한 여부
line_chart.title = "성적표"
# line_chart.title line_chart 객체 대한 제목 결정
line_chart.style = 13
# line_chart.style line_chart 객체에 대한 스타일을 지정 (번호를 통해 지정가능)
line_chart.y_axis.title = "점수"
line_chart.x_axis.title = "번호"
# 각 x, y 열과 행에 대한 제목을 선정

ws.add_chart(line_chart, "E1")
# add_chart() 시트에 chart를 추가하는 함수로 "E1"은 추가하는 위치에 해당.

wb.save("testsampleb.xlsx")
wb.close()