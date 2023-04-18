from openpyxl import Workbook
from random import randint

wb = Workbook()

w1 = wb.create_sheet("number1", 0)

for x in range(1, 100):
    for y in range(1, 100):
        w1.cell(row=x, column=y, value=randint(0, 100))

wb.save("testsampleb.xlsx")
wb.close()